"""XLSX/CSV extractor — uses openpyxl for .xlsx and csv stdlib for .csv."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass


@dataclass(frozen=True)
class XlsxResult:
    full_text: str
    sheet_names: tuple[str, ...]
    row_count: int
    tables_as_markdown: str
    language: str | None


def extract_xlsx(raw: bytes, *, filename: str = "") -> XlsxResult:
    """Extract sheets and tabular data from a .xlsx or .csv byte payload.

    Inputs:  raw bytes of a .xlsx or .csv file; optional filename hint for
             format detection when bytes alone are ambiguous.
    Outputs: XlsxResult with concatenated text, sheet names, total row count,
             all sheets rendered as Markdown tables, and detected language.
    Failures: raises ValueError if openpyxl cannot open the bytes;
              csv.Error if the CSV payload is malformed;
              ImportError if openpyxl is not installed.
    """
    if _looks_like_csv(raw, filename):
        return _extract_csv(raw)
    return _extract_xlsx_bytes(raw)


# ---------------------------------------------------------------------------
# Internal per-format extractors
# ---------------------------------------------------------------------------

def _extract_xlsx_bytes(raw: bytes) -> XlsxResult:
    """Parse an xlsx workbook from bytes."""
    import openpyxl  # deferred import

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_names: list[str] = list(wb.sheetnames)
    all_markdown: list[str] = []
    total_rows = 0

    for name in sheet_names:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        total_rows += len(rows)
        md = _rows_to_markdown(name, rows)
        all_markdown.append(md)

    tables_as_markdown = "\n\n".join(all_markdown)
    full_text = tables_as_markdown
    language = _detect_language_hint(full_text)

    return XlsxResult(
        full_text=full_text,
        sheet_names=tuple(sheet_names),
        row_count=total_rows,
        tables_as_markdown=tables_as_markdown,
        language=language,
    )


def _extract_csv(raw: bytes) -> XlsxResult:
    """Parse a CSV payload from bytes."""
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    rows: list[tuple[object, ...]] = [tuple(row) for row in reader]
    if not rows:
        return XlsxResult(
            full_text="",
            sheet_names=("Sheet1",),
            row_count=0,
            tables_as_markdown="",
            language=None,
        )

    md = _rows_to_markdown("Sheet1", rows)
    language = _detect_language_hint(md)

    return XlsxResult(
        full_text=md,
        sheet_names=("Sheet1",),
        row_count=len(rows),
        tables_as_markdown=md,
        language=language,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _looks_like_csv(raw: bytes, filename: str) -> bool:
    """Return True if the payload should be treated as CSV."""
    if filename.lower().endswith(".csv"):
        return True
    # Sniff first 512 bytes: if no xlsx PK magic, assume CSV.
    if raw[:4] == b"PK\x03\x04":
        return False
    try:
        sample = raw[:512].decode("utf-8", errors="replace")
        csv.Sniffer().sniff(sample, delimiters=",;\t")
        return True
    except csv.Error:
        return False


def _rows_to_markdown(sheet_name: str, rows: list[tuple[object, ...]]) -> str:
    """Render a list of row tuples as a Markdown table with a section header."""
    if not rows:
        return f"## {sheet_name}\n\n_(empty)_"

    def cell(v: object) -> str:
        return str(v).replace("|", "\\|").replace("\n", " ") if v is not None else ""

    header = rows[0]
    separator = ["---"] * len(header)
    header_line = "| " + " | ".join(cell(c) for c in header) + " |"
    sep_line = "| " + " | ".join(separator) + " |"
    data_lines = [
        "| " + " | ".join(cell(c) for c in row) + " |"
        for row in rows[1:]
    ]
    table = "\n".join([header_line, sep_line] + data_lines)
    return f"## {sheet_name}\n\n{table}"


_LANG_HINTS: dict[str, frozenset[str]] = {
    "en": frozenset({"the", "and", "of", "to", "in", "is", "it", "that"}),
    "es": frozenset({"de", "la", "el", "en", "que", "y", "los", "las", "un"}),
    "fr": frozenset({"de", "la", "le", "les", "et", "en", "un", "une", "du"}),
    "de": frozenset({"der", "die", "das", "und", "in", "zu", "den", "ist"}),
}


def _detect_language_hint(text: str) -> str | None:
    """Return ISO-639-1 code for the most likely language, or None if uncertain."""
    words = re.findall(r"\b[a-zA-Z]{2,}\b", text.lower())
    if not words:
        return None
    sample = set(words[:500])
    scores = {lang: len(sample & hints) for lang, hints in _LANG_HINTS.items()}
    best_lang, best_score = max(scores.items(), key=lambda x: x[1])
    return best_lang if best_score >= 3 else None
